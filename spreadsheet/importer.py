"""
Excel Importer
==============
Reads an .xlsx file with openpyxl and populates the DB:
  - Parses cell values and formulas
  - Converts Excel formulas → Python expressions
  - Builds the dependency graph (CellDependency)

Usage:
    from spreadsheet.importer import import_excel
    sheet = import_excel('path/to/file.xlsx', sheet_index=0, sheet_name='Бюджет')
"""

import re
from .models import Sheet, Cell, CellDependency
from .formula_engine import formula_to_python, extract_cell_refs


def _get_cell_format(number_format: str):
    """Guess format_type from Excel number format string.
    Only applies currency/percent/number formatting if explicitly present in the format string.
    Plain numbers get no special format_type so they display as-is.
    """
    if not number_format or number_format == 'General':
        return '', 0
    nf = number_format.lower()

    # Currency: only if an explicit currency symbol is present
    currency_symbols = ['₸', '$', '€', '£', '¥', '"руб"', 'rub', '[$₸', '[$€', '[$£', '[$¥']
    if any(sym in nf for sym in currency_symbols):
        # Count decimal places from format string
        dec = 2
        if '.00' in nf: dec = 2
        elif '.0' in nf and '.00' not in nf: dec = 1
        elif '0.' not in nf: dec = 0
        return 'currency', dec

    # Percent: only if % is present
    if '%' in nf:
        dec = 1 if '.0' in nf else 0
        return 'percent', dec

    # Explicit decimal number format (e.g. 0.00, #,##0.00) — show as number
    if re.search(r'0\.0+', nf):
        dec = len(re.search(r'0\.(0+)', nf).group(1)) if re.search(r'0\.(0+)', nf) else 2
        return 'number', dec

    # Everything else (plain integers like #,##0 or General) — no special format
    return '', 0


def _color_to_hex(color_obj):
    """Convert openpyxl Color to hex string."""
    if color_obj is None:
        return ''
    try:
        if hasattr(color_obj, 'rgb') and color_obj.rgb and color_obj.rgb != '00000000':
            rgb = color_obj.rgb
            # Skip fully transparent or default
            if rgb.startswith('FF') or rgb.startswith('ff'):
                return '#' + rgb[2:8]
            return '#' + rgb[2:8] if len(rgb) == 8 else ''
    except Exception:
        pass
    return ''


def import_excel(filepath: str, sheet_index: int = 0, sheet_name: str = None) -> Sheet:
    """
    Import an Excel sheet into the database.
    Returns the created Sheet instance.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = load_workbook(filepath, data_only=False)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[sheet_index]

    # Determine actual used range
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Create or replace sheet
    name = sheet_name or ws.title
    sheet, _ = Sheet.objects.get_or_create(name=name)
    # Clear existing data for re-import
    sheet.cells.all().delete()
    sheet.dependencies.all().delete()

    cells_to_create = []
    formula_cells = []

    # Track merged cells
    merged_ranges = {}
    for merge_range in ws.merged_cells.ranges:
        min_row, min_col = merge_range.min_row - 1, merge_range.min_col - 1
        max_row_m, max_col_m = merge_range.max_row - 1, merge_range.max_col - 1
        merged_ranges[(min_row, min_col)] = {
            'row_span': max_row_m - min_row + 1,
            'col_span': max_col_m - min_col + 1,
        }
        # Mark spanned cells to skip
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                if not (r == min_row and c == min_col):
                    merged_ranges[(r, c)] = 'skip'

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col)):
        for col_idx, excel_cell in enumerate(row):
            rc = (row_idx, col_idx)

            if merged_ranges.get(rc) == 'skip':
                continue

            merge_info = merged_ranges.get(rc, {})

            raw = excel_cell.value
            formula = None
            py_formula = None
            cell_type = 'input'
            is_editable = True

            if isinstance(raw, str) and raw.startswith('='):
                formula = raw
                py_formula = formula_to_python(raw[1:])
                cell_type = 'formula'
                is_editable = False
                value = '0'  # Will be recalculated
                formula_cells.append((row_idx, col_idx, raw))
            elif raw is None:
                value = ''
                cell_type = 'label'
                is_editable = False
            elif isinstance(raw, (int, float)):
                value = str(raw)
                is_editable = True
            elif hasattr(raw, 'hour') and not hasattr(raw, 'year'):
                # datetime.time — in Excel a zero date shows as time(0,0)
                # treat as 0 (no date set)
                from datetime import time as _time
                value = '0' if raw == _time(0, 0) else str(raw.hour * 3600 / 86400)
                is_editable = True
            elif isinstance(raw, str):
                value = raw
                # Detect headers/labels: non-numeric strings in first few rows or columns
                if col_idx < 3 or row_idx < 3:
                    cell_type = 'header' if row_idx < 2 else 'label'
                    is_editable = False
            else:
                value = str(raw) if raw is not None else ''

            # Styling
            font = excel_cell.font
            fill = excel_cell.fill
            fmt_type, dec = _get_cell_format(excel_cell.number_format)

            bold = font.bold if font else False
            italic = font.italic if font else False
            text_color = ''
            bg_color = ''

            if font and font.color:
                text_color = _color_to_hex(font.color)
            if fill and fill.fgColor:
                bg_color = _color_to_hex(fill.fgColor)

            cell_obj = Cell(
                sheet=sheet,
                row=row_idx,
                col=col_idx,
                value=value,
                raw_value=str(raw) if raw is not None else '',
                formula=formula,
                python_formula=py_formula,
                cell_type=cell_type,
                is_editable=is_editable,
                format_type=fmt_type,
                decimal_places=dec,
                row_span=merge_info.get('row_span', 1),
                col_span=merge_info.get('col_span', 1),
                bold=bool(bold),
                italic=bool(italic),
                text_color=text_color,
                bg_color=bg_color,
                number_format=excel_cell.number_format or '',
            )
            cells_to_create.append(cell_obj)

    Cell.objects.bulk_create(cells_to_create, ignore_conflicts=True)

    # Build dependency graph
    deps_to_create = []
    for (row_idx, col_idx, formula) in formula_cells:
        refs = extract_cell_refs(formula)
        for (r, c) in refs:
            deps_to_create.append(CellDependency(
                sheet=sheet,
                source_row=r,
                source_col=c,
                target_row=row_idx,
                target_col=col_idx,
            ))

    if deps_to_create:
        CellDependency.objects.bulk_create(deps_to_create, ignore_conflicts=True)

    # Initial full recalculation
    _recalc_all(sheet)

    return sheet


def _recalc_all(sheet):
    """
    Full recalculation after import.
    Uses topological sort so formulas that depend on other formulas
    are always evaluated after their dependencies.
    """
    from .formula_engine import make_eval_context, extract_cell_refs
    import logging
    from collections import defaultdict, deque
    logger = logging.getLogger(__name__)

    cells = list(sheet.cells.all())
    cell_map = {(c.row, c.col): c for c in cells}

    # Seed cell_values with all non-formula cells
    cell_values = {}
    for c in cells:
        if not c.python_formula:
            try:
                cell_values[(c.row, c.col)] = float(c.value) if c.value not in (None, '') else 0
            except (TypeError, ValueError):
                cell_values[(c.row, c.col)] = c.value or 0

    # Override with raw_value for datetime cells (they get stored as str in .value)
    from datetime import datetime as _dt, time as _time
    for c in cells:
        if not c.python_formula and c.raw_value:
            rv = c.raw_value
            # time(0,0) was stored as "00:00:00" or "0:00:00" — means "no date" = 0
            if rv in ('0:00:00', '00:00:00'):
                cell_values[(c.row, c.col)] = 0
                continue
            # openpyxl stores datetime repr like "2024-08-12 00:00:00"
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                try:
                    cell_values[(c.row, c.col)] = _dt.strptime(rv, fmt)
                    break
                except ValueError:
                    pass

    formula_cells = [c for c in cells if c.python_formula]
    if not formula_cells:
        return

    # Build dependency graph between formula cells
    # adj[A] = list of formula cells that depend on A
    formula_coords = {(c.row, c.col) for c in formula_cells}
    in_degree = {(c.row, c.col): 0 for c in formula_cells}
    dependents = defaultdict(list)  # coord -> list of formula coords that need it

    for c in formula_cells:
        refs = extract_cell_refs(c.formula or '')
        for ref in refs:
            if ref in formula_coords:
                dependents[ref].append((c.row, c.col))
                in_degree[(c.row, c.col)] += 1

    # Kahn's algorithm — topological sort
    queue = deque(coord for coord, deg in in_degree.items() if deg == 0)
    topo_order = []
    while queue:
        coord = queue.popleft()
        topo_order.append(coord)
        for dep_coord in dependents[coord]:
            in_degree[dep_coord] -= 1
            if in_degree[dep_coord] == 0:
                queue.append(dep_coord)

    # If there are cycles, append remaining cells at the end
    remaining = [coord for coord, deg in in_degree.items() if deg > 0]
    topo_order.extend(remaining)

    # Evaluate in topological order — each formula sees already-computed dependencies
    to_save = []
    for coord in topo_order:
        c = cell_map.get(coord)
        if not c or not c.python_formula:
            continue
        ctx = make_eval_context(cell_values)
        try:
            result = eval(c.python_formula, {"__builtins__": {}}, ctx)
            if isinstance(result, (int, float)):
                dp = c.decimal_places if c.decimal_places is not None else 10
                result = round(float(result), dp)
            new_val = str(result) if result is not None else '0'
            cell_values[coord] = result
            if new_val != c.value:
                c.value = new_val
                to_save.append(c)
        except Exception as e:
            logger.warning(
                f"Formula error at {coord} formula=[{c.formula}] "
                f"python=[{c.python_formula}]: {e}"
            )
            cell_values[coord] = 0

    if to_save:
        Cell.objects.bulk_update(to_save, ['value'])
