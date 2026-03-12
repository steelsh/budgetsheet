"""
Apply Data
==========
Берёт существующий шаблон из БД (например, импортированный файл 2024)
и накладывает на него входные данные из нового Excel-файла (2025).

Логика:
  - Проходим по всем ячейкам нового файла
  - Если ячейка НЕ пустая И в шаблоне на той же позиции стоит входная ячейка (is_editable=True)
    → подставляем новое значение
  - Если в шаблоне формула или метка → не трогаем
  - Создаём НОВЫЙ лист (шаблон остаётся нетронутым)
  - Пересчитываем все формулы с новыми данными
"""

import copy
from .models import Sheet, Cell, CellDependency
from .formula_engine import formula_to_python, extract_cell_refs, recalculate_sheet, format_value


def apply_data_file(template_sheet: Sheet, data_path: str,
                    data_sheet_index: int = 0,
                    new_sheet_name: str = None) -> dict:
    """
    Creates a new Sheet based on template_sheet, with input values
    replaced by values from data_path Excel file.

    Returns:
        {
            'sheet': Sheet instance,
            'replaced': int,   # how many input cells were updated
            'skipped': int,    # cells in data file that had no matching input in template
            'kept': int,       # input cells kept from template (data file was empty there)
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = load_workbook(data_path, data_only=False)
    ws = wb.worksheets[data_sheet_index]

    # Build lookup of new data: (row, col) -> value  (only non-empty, non-formula cells)
    new_data = {}
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            val = cell.value
            if val is None or str(val).strip() == '':
                continue
            if isinstance(val, str) and val.strip().startswith('='):
                continue  # ignore formulas in data file
            new_data[(row_idx, col_idx)] = val

    # Load all template cells
    template_cells = list(template_sheet.cells.all().order_by('row', 'col'))
    template_deps = list(template_sheet.dependencies.all())

    # Create new sheet
    if not new_sheet_name:
        import re
        base = template_sheet.name
        # Try to bump year: "Бюджет 2024" -> "Бюджет 2025"
        match = re.search(r'(\d{4})', base)
        if match:
            year = int(match.group(1))
            new_sheet_name = base.replace(str(year), str(year + 1))
        else:
            new_sheet_name = base + ' (новый)'

    # Remove existing sheet with same name if present
    Sheet.objects.filter(name=new_sheet_name).delete()

    new_sheet = Sheet.objects.create(name=new_sheet_name)

    stats = {'replaced': 0, 'skipped': 0, 'kept': 0}
    new_cells = []

    for tc in template_cells:
        pos = (tc.row, tc.col)
        data_val = new_data.get(pos)

        if tc.is_editable and data_val is not None:
            # Replace with new value from data file
            clean = str(data_val).strip().replace(',', '.')
            try:
                float(clean)
                final_value = clean
            except ValueError:
                final_value = str(data_val)
            stats['replaced'] += 1
        elif tc.is_editable and data_val is None:
            # Keep template value
            final_value = tc.value
            stats['kept'] += 1
        else:
            # Formula or label — copy as-is, recalculate later
            final_value = tc.value if not tc.python_formula else '0'

        new_cells.append(Cell(
            sheet=new_sheet,
            row=tc.row,
            col=tc.col,
            value=final_value,
            raw_value=tc.raw_value,
            formula=tc.formula,
            python_formula=tc.python_formula,
            cell_type=tc.cell_type,
            is_editable=tc.is_editable,
            format_type=tc.format_type,
            decimal_places=tc.decimal_places,
            row_span=tc.row_span,
            col_span=tc.col_span,
            bold=tc.bold,
            italic=tc.italic,
            bg_color=tc.bg_color,
            text_color=tc.text_color,
            number_format=tc.number_format,
        ))

    Cell.objects.bulk_create(new_cells)

    # Copy dependency graph
    new_deps = [
        CellDependency(
            sheet=new_sheet,
            source_row=d.source_row,
            source_col=d.source_col,
            target_row=d.target_row,
            target_col=d.target_col,
        )
        for d in template_deps
    ]
    if new_deps:
        CellDependency.objects.bulk_create(new_deps, ignore_conflicts=True)

    # Count skipped: data cells that had no matching input in template
    template_input_positions = {(c.row, c.col) for c in template_cells if c.is_editable}
    stats['skipped'] = sum(
        1 for pos in new_data if pos not in template_input_positions
    )

    # Full recalculation with new values
    changed, _ = recalculate_sheet(new_sheet)
    if changed:
        formula_cells = list(new_sheet.cells.filter(cell_type='formula'))
        cell_map = {(c.row, c.col): c for c in formula_cells}
        to_save = []
        for (r, c), val in changed.items():
            cell = cell_map.get((r, c))
            if cell:
                cell.value = val
                to_save.append(cell)
        if to_save:
            Cell.objects.bulk_update(to_save, ['value'])

    return {'sheet': new_sheet, **stats}
