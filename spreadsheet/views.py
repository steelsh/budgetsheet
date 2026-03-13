import json
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User

from .models import Sheet, Cell, ChangeHistory, SheetSnapshot, CellDependency
from .formula_engine import recalculate_dependents, format_value, formula_display


# ─── Auth ───────────────────────────────────────────────────────────────

def login_page(request):
    if request.user.is_authenticated:
        return redirect('/')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect(request.GET.get('next', '/'))
        error = 'Неверный логин или пароль'
    return render(request, 'spreadsheet/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('/login/')


def register_page(request):
    if request.user.is_authenticated:
        return redirect('/')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')
        if not username or not password:
            error = 'Заполните все поля'
        elif password != password2:
            error = 'Пароли не совпадают'
        elif User.objects.filter(username=username).exists():
            error = 'Пользователь уже существует'
        else:
            user = User.objects.create_user(username=username, password=password)
            login(request, user)
            return redirect('/')
    return render(request, 'spreadsheet/register.html', {'error': error})


# ─── Main views ─────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def index(request):
    sheets = Sheet.objects.all()
    sheet = sheets.first()
    if not sheet:
        from .demo_data import seed_demo_data
        sheet = seed_demo_data()
    return render(request, 'spreadsheet/index.html', {'sheet': sheet, 'sheets': sheets})


@login_required(login_url='/login/')
def sheet_view(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    sheets = Sheet.objects.all()
    return render(request, 'spreadsheet/index.html', {'sheet': sheet, 'sheets': sheets})


@login_required(login_url='/login/')
def get_sheet_data(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    cells = list(sheet.cells.all().order_by('row', 'col'))
    max_row = max((c.row for c in cells), default=0)
    max_col = max((c.col for c in cells), default=0)
    cell_data = []
    for c in cells:
        cell_data.append({
            'id': c.id,
            'row': c.row,
            'col': c.col,
            'value': c.value,
            'formatted': format_value(c.value, c.format_type, c.decimal_places),
            'type': c.cell_type,
            'editable': c.is_editable,
            'formula': formula_display(c.formula) if c.formula else '',
            'bold': c.bold,
            'italic': c.italic,
            'bg_color': c.bg_color,
            'text_color': c.text_color,
            'format_type': c.format_type,
            'col_span': c.col_span,
            'row_span': c.row_span,
            'comment': c.comment or '',
        })
    return JsonResponse({
        'sheet_id': sheet.id,
        'sheet_name': sheet.name,
        'max_row': max_row,
        'max_col': max_col,
        'cells': cell_data,
    })


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def update_cell(request, sheet_id):
    try:
        body = json.loads(request.body)
        row = int(body['row'])
        col = int(body['col'])
        new_value = str(body['value']).strip()
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'error': f'Неверный запрос: {e}'}, status=400)

    sheet = get_object_or_404(Sheet, id=sheet_id)
    try:
        cell = Cell.objects.get(sheet=sheet, row=row, col=col)
    except Cell.DoesNotExist:
        return JsonResponse({'error': 'Ячейка не найдена'}, status=404)

    if not cell.is_editable:
        return JsonResponse({'error': 'Ячейка не редактируема'}, status=403)

    clean = new_value.replace(' ', '').replace('₸', '').replace('%', '').replace(',', '.')
    try:
        float(clean)
        new_value = clean
    except ValueError:
        pass

    old_value = cell.value
    Cell.objects.filter(pk=cell.pk).update(value=new_value)

    try:
        ChangeHistory.objects.create(
            cell=cell,
            old_value=old_value,
            new_value=new_value,
            changed_by=request.user.username if request.user.is_authenticated else request.META.get('REMOTE_ADDR', 'unknown'),
        )
    except Exception:
        pass

    updates = recalculate_dependents(sheet, row, col, new_value)
    edited_formatted = format_value(new_value, cell.format_type, cell.decimal_places)
    return JsonResponse({
        'success': True,
        'edited_cell': {'row': row, 'col': col, 'value': new_value, 'formatted': edited_formatted},
        'updates': updates,
        'updates_count': len(updates),
    })


# ─── Comments ────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
@require_http_methods(["POST"])
def update_comment(request, sheet_id):
    try:
        body = json.loads(request.body)
        row = int(body['row'])
        col = int(body['col'])
        comment = str(body.get('comment', '')).strip()
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'error': str(e)}, status=400)

    sheet = get_object_or_404(Sheet, id=sheet_id)
    try:
        cell = Cell.objects.get(sheet=sheet, row=row, col=col)
    except Cell.DoesNotExist:
        return JsonResponse({'error': 'Ячейка не найдена'}, status=404)

    Cell.objects.filter(pk=cell.pk).update(comment=comment or None)
    return JsonResponse({'success': True, 'comment': comment})


# ─── Search ──────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def search_cells(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse({'results': []})

    cells = sheet.cells.all()
    results = []
    q_lower = q.lower()
    for c in cells:
        val = str(c.value or '').lower()
        formula = str(c.formula or '').lower()
        comment = str(c.comment or '').lower()
        if q_lower in val or q_lower in formula or q_lower in comment:
            results.append({
                'row': c.row,
                'col': c.col,
                'ref': c.excel_ref,
                'value': c.value or '',
                'formatted': format_value(c.value, c.format_type, c.decimal_places),
                'formula': c.formula or '',
                'comment': c.comment or '',
                'type': c.cell_type,
            })
    return JsonResponse({'results': results, 'count': len(results)})


# ─── History ─────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def cell_history(request, sheet_id, row, col):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    try:
        cell = Cell.objects.get(sheet=sheet, row=row, col=col)
    except Cell.DoesNotExist:
        return JsonResponse({'history': [], 'cell_ref': ''})
    history = list(cell.history.order_by('-changed_at').values(
        'old_value', 'new_value', 'changed_at', 'changed_by'
    )[:50])
    for h in history:
        h['changed_at'] = h['changed_at'].strftime('%d.%m.%Y %H:%M:%S')
    return JsonResponse({'history': history, 'cell_ref': cell.excel_ref})


# ─── Snapshots (версионирование) ─────────────────────────────────────────

@login_required(login_url='/login/')
@require_http_methods(["POST"])
def create_snapshot(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    try:
        body = json.loads(request.body)
        note = str(body.get('note', '')).strip()
    except Exception:
        note = ''

    cells_data = []
    for c in sheet.cells.all():
        cells_data.append({
            'row': c.row, 'col': c.col, 'value': c.value,
            'formula': c.formula, 'python_formula': c.python_formula,
            'cell_type': c.cell_type, 'is_editable': c.is_editable,
            'format_type': c.format_type, 'decimal_places': c.decimal_places,
            'bold': c.bold, 'italic': c.italic,
            'bg_color': c.bg_color, 'text_color': c.text_color,
            'row_span': c.row_span, 'col_span': c.col_span,
            'comment': c.comment,
        })

    snap = SheetSnapshot.objects.create(
        sheet=sheet,
        name=f"{sheet.name} — {__import__('datetime').datetime.now().strftime('%d.%m.%Y %H:%M')}",
        note=note,
        data={'cells': cells_data},
        created_by=request.user.username if request.user.is_authenticated else 'unknown',
    )
    return JsonResponse({'success': True, 'snapshot_id': snap.id, 'name': snap.name, 'created_at': snap.created_at.strftime('%d.%m.%Y %H:%M')})


@login_required(login_url='/login/')
def list_snapshots(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    snaps = list(sheet.snapshots.values('id', 'name', 'note', 'created_at', 'created_by'))
    for s in snaps:
        s['created_at'] = s['created_at'].strftime('%d.%m.%Y %H:%M')
    return JsonResponse({'snapshots': snaps})


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def restore_snapshot(request, sheet_id, snapshot_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)
    snap = get_object_or_404(SheetSnapshot, id=snapshot_id, sheet=sheet)

    # Save current state as snapshot before restoring
    cells_now = []
    for c in sheet.cells.all():
        cells_now.append({'row': c.row, 'col': c.col, 'value': c.value,
            'formula': c.formula, 'python_formula': c.python_formula,
            'cell_type': c.cell_type, 'is_editable': c.is_editable,
            'format_type': c.format_type, 'decimal_places': c.decimal_places,
            'bold': c.bold, 'italic': c.italic,
            'bg_color': c.bg_color, 'text_color': c.text_color,
            'row_span': c.row_span, 'col_span': c.col_span, 'comment': c.comment})
    SheetSnapshot.objects.create(
        sheet=sheet,
        name=f"Авто-сохранение перед откатом {__import__('datetime').datetime.now().strftime('%d.%m.%Y %H:%M')}",
        note='Создано автоматически перед откатом',
        data={'cells': cells_now},
        created_by=request.user.username if request.user.is_authenticated else 'unknown',
    )

    # Restore
    cells_data = snap.data.get('cells', [])
    sheet.cells.all().delete()
    new_cells = []
    for cd in cells_data:
        new_cells.append(Cell(
            sheet=sheet,
            row=cd['row'], col=cd['col'], value=cd.get('value'),
            formula=cd.get('formula'), python_formula=cd.get('python_formula'),
            cell_type=cd.get('cell_type', 'input'),
            is_editable=cd.get('is_editable', False),
            format_type=cd.get('format_type', ''),
            decimal_places=cd.get('decimal_places', 2),
            bold=cd.get('bold', False), italic=cd.get('italic', False),
            bg_color=cd.get('bg_color', ''), text_color=cd.get('text_color', ''),
            row_span=cd.get('row_span', 1), col_span=cd.get('col_span', 1),
            comment=cd.get('comment'),
        ))
    Cell.objects.bulk_create(new_cells)
    return JsonResponse({'success': True, 'message': f'Откат к версии "{snap.name}" выполнен'})


# ─── Compare sheets ──────────────────────────────────────────────────────

@login_required(login_url='/login/')
def compare_sheets(request):
    sheets = Sheet.objects.all()
    return render(request, 'spreadsheet/compare.html', {'sheets': sheets})


@login_required(login_url='/login/')
def compare_api(request):
    a_id = request.GET.get('a')
    b_id = request.GET.get('b')
    if not a_id or not b_id:
        return JsonResponse({'error': 'Укажите два листа'}, status=400)

    sheet_a = get_object_or_404(Sheet, id=a_id)
    sheet_b = get_object_or_404(Sheet, id=b_id)

    cells_a = {(c.row, c.col): c for c in sheet_a.cells.all()}
    cells_b = {(c.row, c.col): c for c in sheet_b.cells.all()}
    all_positions = set(cells_a.keys()) | set(cells_b.keys())

    diffs = []
    for pos in sorted(all_positions):
        ca = cells_a.get(pos)
        cb = cells_b.get(pos)
        val_a = ca.value if ca else None
        val_b = cb.value if cb else None
        if str(val_a or '') != str(val_b or ''):
            row, col = pos
            ref = f"{chr(65 + col % 26)}{row + 1}"
            diffs.append({
                'ref': ref, 'row': row, 'col': col,
                'value_a': format_value(val_a, ca.format_type if ca else '', ca.decimal_places if ca else 2) if val_a else '—',
                'value_b': format_value(val_b, cb.format_type if cb else '', cb.decimal_places if cb else 2) if val_b else '—',
                'raw_a': val_a or '',
                'raw_b': val_b or '',
                'type': 'changed' if ca and cb else ('only_a' if ca else 'only_b'),
            })

    return JsonResponse({
        'sheet_a': sheet_a.name,
        'sheet_b': sheet_b.name,
        'diffs': diffs,
        'diff_count': len(diffs),
    })


# ─── Export to Excel ─────────────────────────────────────────────────────

@login_required(login_url='/login/')
def export_excel(request, sheet_id):
    sheet = get_object_or_404(Sheet, id=sheet_id)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl не установлен', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet.name[:31]

    cells = sheet.cells.all().order_by('row', 'col')
    for c in cells:
        excel_row = c.row + 1
        excel_col = c.col + 1
        cell = ws.cell(row=excel_row, column=excel_col)

        # Value: use formula if available, else calculated value
        if c.formula:
            cell.value = c.formula
        else:
            raw = c.value
            if raw is not None and raw != '':
                try:
                    cell.value = float(raw)
                except (ValueError, TypeError):
                    cell.value = raw
            else:
                cell.value = None

        # Styling
        font_kwargs = {}
        if c.bold: font_kwargs['bold'] = True
        if c.italic: font_kwargs['italic'] = True
        if c.text_color and c.text_color.startswith('#'):
            font_kwargs['color'] = c.text_color[1:].upper().zfill(8)
        if font_kwargs:
            cell.font = Font(**font_kwargs)

        if c.bg_color and c.bg_color.startswith('#'):
            fill_color = c.bg_color[1:].upper().zfill(6)
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)

        if c.number_format:
            cell.number_format = c.number_format

        if c.col_span > 1 or c.row_span > 1:
            end_row = excel_row + c.row_span - 1
            end_col = excel_col + c.col_span - 1
            ws.merge_cells(
                start_row=excel_row, start_column=excel_col,
                end_row=end_row, end_column=end_col,
            )

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ''))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sheet.name}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Import ──────────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def excel_sheets_list(request):
    """Return list of sheet names in uploaded Excel file (for sheet picker)"""
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'Файл не выбран'}, status=400)
    uploaded = request.FILES['file']
    try:
        from openpyxl import load_workbook
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            for chunk in uploaded.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        os.unlink(tmp_path)
        return JsonResponse({'sheets': sheet_names})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def import_excel_view(request, sheet_id=None):
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'Файл не выбран'}, status=400)
    uploaded = request.FILES['file']
    if not uploaded.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': 'Поддерживается только .xlsx'}, status=400)

    import tempfile, os
    from .importer import import_excel

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        sheet_name = request.POST.get('sheet_name', uploaded.name.replace('.xlsx', ''))
        sheet_index = int(request.POST.get('sheet_index', 0))
        sheet = import_excel(tmp_path, sheet_index=sheet_index, sheet_name=sheet_name)
        return JsonResponse({
            'success': True,
            'sheet_id': sheet.id,
            'sheet_name': sheet.name,
            'message': f'Импортировано: {sheet.cells.count()} ячеек',
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        os.unlink(tmp_path)


# ─── Other pages ─────────────────────────────────────────────────────────

@login_required(login_url='/login/')
def apply_data_page(request):
    sheets = Sheet.objects.all()
    return render(request, 'spreadsheet/apply_data.html', {'sheets': sheets})


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def apply_data_view(request):
    if 'data_file' not in request.FILES:
        return JsonResponse({'error': 'Файл с данными не выбран'}, status=400)
    template_id = request.POST.get('template_id')
    if not template_id:
        return JsonResponse({'error': 'Не выбран шаблон'}, status=400)

    template_sheet = get_object_or_404(Sheet, id=template_id)
    uploaded = request.FILES['data_file']
    new_sheet_name = request.POST.get('new_sheet_name', '').strip() or None
    data_sheet_index = int(request.POST.get('data_sheet_index', 0))

    import tempfile, os
    from .apply_data import apply_data_file

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        result = apply_data_file(
            template_sheet=template_sheet,
            data_path=tmp_path,
            data_sheet_index=data_sheet_index,
            new_sheet_name=new_sheet_name,
        )
        sheet = result['sheet']
        return JsonResponse({
            'success': True,
            'sheet_id': sheet.id,
            'sheet_name': sheet.name,
            'replaced': result['replaced'],
            'kept': result['kept'],
            'skipped': result['skipped'],
            'message': f"Готово: подставлено {result['replaced']} новых значений, сохранено {result['kept']}, пропущено {result['skipped']}",
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        os.unlink(tmp_path)


@login_required(login_url='/login/')
def merge_page(request):
    sheets = Sheet.objects.all()
    return render(request, 'spreadsheet/merge.html', {'sheets': sheets})


@login_required(login_url='/login/')
@require_http_methods(["POST"])
def merge_files(request):
    if 'base_file' not in request.FILES or 'data_file' not in request.FILES:
        return JsonResponse({'error': 'Нужно загрузить оба файла'}, status=400)

    import tempfile, os
    from .merger import merge_excel_files

    base_file = request.FILES['base_file']
    data_file = request.FILES['data_file']
    sheet_name = request.POST.get('sheet_name', '').strip() or f"Merged {data_file.name.replace('.xlsx','')}"

    def save_tmp(f):
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        for chunk in f.chunks():
            tmp.write(chunk)
        tmp.close()
        return tmp.name

    base_path = save_tmp(base_file)
    data_path = save_tmp(data_file)

    try:
        base_sheet_index = int(request.POST.get('base_sheet_index', 0))
        data_sheet_index = int(request.POST.get('data_sheet_index', 0))
        sheet = merge_excel_files(
            base_path, data_path,
            base_sheet_index=base_sheet_index,
            data_sheet_index=data_sheet_index,
            sheet_name=sheet_name,
        )
        return JsonResponse({'success': True, 'sheet_id': sheet.id, 'sheet_name': sheet.name,
                             'message': f'Объединено: {sheet.cells.count()} ячеек'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        os.unlink(base_path)
        os.unlink(data_path)


def seed_demo(request):
    from .demo_data import seed_demo_data
    seed_demo_data()
    return redirect('/')
