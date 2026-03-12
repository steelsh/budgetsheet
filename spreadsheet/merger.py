"""
Excel Merger
============
Combines two Excel files:
  - BASE file (e.g. 2024): provides formulas, structure, labels, styling
  - DATA file (e.g. 2025): provides new input values

Rules per cell:
  1. If BASE cell has a formula (starts with =)  → use formula from BASE, ignore DATA
  2. If BASE cell is a label/header (text)        → keep from BASE as-is
  3. If BASE cell is a plain number AND DATA has a value at same position → use DATA value
  4. If BASE cell is a plain number AND DATA is empty → keep BASE value
  5. Empty cells → empty

Result: a new Sheet in the DB ready to use.
"""

from .models import Sheet, Cell, CellDependency
from .formula_engine import formula_to_python, extract_cell_refs


def _is_formula(val):
    return isinstance(val, str) and val.strip().startswith('=')


def _is_numeric(val):
    if val is None:
        return False
    try:
        float(str(val).replace(',', '.'))
        return True
    except (ValueError, TypeError):
        return False


def _color_to_hex(color_obj):
    if color_obj is None:
        return ''
    try:
        if hasattr(color_obj, 'rgb') and color_obj.rgb and color_obj.rgb != '00000000':
            rgb = color_obj.rgb
            if len(rgb) == 8:
                return '#' + rgb[2:8]
    except Exception:
        pass
    return ''


def _get_format(number_format):
    if not number_format:
        return '', 2
    nf = number_format.lower()
    if any(c in nf for c in ['₸', '$', '€', '#,##0']):
        return 'currency', 2
    if '%' in nf:
        return 'percent', 1
    if '0.00' in nf or '#,##' in nf:
        return 'number', 2
    return '', 0


def merge_excel_files(base_path: str, data_path: str,
                      base_sheet_index: int = 0,
                      data_sheet_index: int = 0,
                      sheet_name: str = 'Merged') -> Sheet:
    """
    Merge base (formulas+structure) with data (new input values).
    Returns the created/updated Sheet instance.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    # Load both workbooks without evaluating formulas
    wb_base = load_workbook(base_path, data_only=False)
    wb_data = load_workbook(data_path, data_only=False)

    ws_base = wb_base.worksheets[base_sheet_index]
    ws_data = wb_data.worksheets[data_sheet_index]

    # Build data lookup: (row_idx, col_idx) -> raw cell value from data file
    data_map = {}
    for row_idx, row in enumerate(ws_data.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.value is not None and str(cell.value).strip() != '':
                data_map[(row_idx, col_idx)] = cell.value

    # Create or replace sheet in DB
    sheet, _ = Sheet.objects.get_or_create(name=sheet_name)
    sheet.cells.all().delete()
    sheet.dependencies.all().delete()

    max_row = ws_base.max_row or 1
    max_col = ws_base.max_column or 1

    # Track merged cells from base
    merged_ranges = {}
    for merge_range in ws_base.merged_cells.ranges:
        min_row = merge_range.min_row - 1
        min_col = merge_range.min_col - 1
        max_row_m = merge_range.max_row - 1
        max_col_m = merge_range.max_col - 1
        merged_ranges[(min_row, min_col)] = {
            'row_span': max_row_m - min_row + 1,
            'col_span': max_col_m - min_col + 1,
        }
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                if not (r == min_row and c == min_col):
                    merged_ranges[(r, c)] = 'skip'

    cells_to_create = []
    formula_cells = []
    stats = {'formulas': 0, 'data_replaced': 0, 'kept_base': 0, 'labels': 0}

    for row_idx, row in enumerate(ws_base.iter_rows(min_row=1, max_row=max_row, max_col=max_col)):
        for col_idx, base_cell in enumerate(row):
            rc = (row_idx, col_idx)

            if merged_ranges.get(rc) == 'skip':
                continue

            merge_info = merged_ranges.get(rc, {})
            base_val = base_cell.value
            data_val = data_map.get(rc)  # value from data file at same position

            # ── Decide what goes in this cell ──────────────────────────────
            formula = None
            py_formula = None
            cell_type = 'input'
            is_editable = True
            value = ''

            if _is_formula(base_val):
                # Rule 1: Formula from base always wins
                formula = str(base_val)
                py_formula = formula_to_python(formula[1:])
                cell_type = 'formula'
                is_editable = False
                value = '0'
                formula_cells.append((row_idx, col_idx, formula))
                stats['formulas'] += 1

            elif base_val is None or str(base_val).strip() == '':
                # Empty base cell — check if data file has something
                if data_val is not None and not _is_formula(str(data_val)):
                    value = str(data_val)
                    cell_type = 'input'
                    is_editable = True
                else:
                    value = ''
                    cell_type = 'label'
                    is_editable = False

            elif isinstance(base_val, str) and not _is_numeric(base_val):
                # Rule 2: Text label — keep from base
                value = base_val
                cell_type = 'label' if row_idx > 1 else 'header'
                is_editable = False
                stats['labels'] += 1

            elif _is_numeric(base_val):
                # Rule 3/4: Numeric cell — prefer data file value if available
                if data_val is not None and not _is_formula(str(data_val)) and str(data_val).strip() != '':
                    value = str(data_val).replace(',', '.')
                    cell_type = 'input'
                    is_editable = True
                    stats['data_replaced'] += 1
                else:
                    value = str(base_val)
                    cell_type = 'input'
                    is_editable = True
                    stats['kept_base'] += 1
            else:
                value = str(base_val) if base_val is not None else ''
                cell_type = 'input'
                is_editable = True

            # ── Styling from base file ──────────────────────────────────────
            font = base_cell.font
            fill = base_cell.fill
            fmt_type, dec = _get_format(base_cell.number_format)
            bold = bool(font.bold) if font else False
            italic = bool(font.italic) if font else False
            text_color = _color_to_hex(font.color) if font else ''
            bg_color = _color_to_hex(fill.fgColor) if fill else ''

            cells_to_create.append(Cell(
                sheet=sheet,
                row=row_idx,
                col=col_idx,
                value=value,
                raw_value=str(base_val) if base_val is not None else '',
                formula=formula,
                python_formula=py_formula,
                cell_type=cell_type,
                is_editable=is_editable,
                format_type=fmt_type,
                decimal_places=dec,
                row_span=merge_info.get('row_span', 1),
                col_span=merge_info.get('col_span', 1),
                bold=bold,
                italic=italic,
                text_color=text_color,
                bg_color=bg_color,
                number_format=base_cell.number_format or '',
            ))

    Cell.objects.bulk_create(cells_to_create, ignore_conflicts=True)

    # Build dependency graph
    deps = []
    for (row_idx, col_idx, formula) in formula_cells:
        for (ref_r, ref_c) in extract_cell_refs(formula):
            deps.append(CellDependency(
                sheet=sheet,
                source_row=ref_r,
                source_col=ref_c,
                target_row=row_idx,
                target_col=col_idx,
            ))
    if deps:
        CellDependency.objects.bulk_create(deps, ignore_conflicts=True)

    # Full recalculation pass
    from .formula_engine import recalculate_sheet
    changed, _ = recalculate_sheet(sheet)
    if changed:
        all_formula_cells = list(sheet.cells.filter(cell_type='formula'))
        cell_map = {(c.row, c.col): c for c in all_formula_cells}
        to_save = []
        for (r, c), val in changed.items():
            cell = cell_map.get((r, c))
            if cell:
                cell.value = val
                to_save.append(cell)
        if to_save:
            Cell.objects.bulk_update(to_save, ['value'])

    # Store merge stats in sheet name as info (optional)
    print(f"[Merger] Done: {stats['formulas']} formulas from base, "
          f"{stats['data_replaced']} values replaced from data file, "
          f"{stats['kept_base']} base values kept, "
          f"{stats['labels']} labels")

    return sheet
